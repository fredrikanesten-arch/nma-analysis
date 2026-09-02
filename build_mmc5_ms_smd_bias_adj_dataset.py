#!/usr/bin/env python3
"""Build an arm-level SMD dataset from the 'MS SMD bias-adj' sheet in mmc5.xlsx.

The output schema matches combined_long_mean_change_dataset_ms.csv columns.
"""

from __future__ import annotations

import argparse
import csv
import math
from dataclasses import dataclass
from statistics import NormalDist
from typing import Iterable

import openpyxl


OUTPUT_COLUMNS = [
    "studyid",
    "na",
    "arm",
    "treatment",
    "n",
    "mean_change",
    "sd_change",
    "source",
    "y_baseline",
    "sd_baseline",
    "y_followup",
    "sd_followup",
    "r_used",
    "responders",
    "p_response",
    "q",
    "cutoff",
]


@dataclass
class ArmRow:
    studyid: str
    na: int
    arm: int
    treatment: int
    n: int
    mean_change: float
    sd_change: float
    source: str
    y_baseline: float | None = None
    sd_baseline: float | None = None
    y_followup: float | None = None
    sd_followup: float | None = None
    r_used: float | None = None
    responders: int | None = None
    p_response: float | None = None
    q: float | None = None
    cutoff: float | None = None

    def to_csv_row(self) -> dict[str, str | int | float]:
        def v(x: object) -> str | int | float:
            return "NA" if x is None else x

        return {
            "studyid": self.studyid,
            "na": self.na,
            "arm": self.arm,
            "treatment": self.treatment,
            "n": self.n,
            "mean_change": self.mean_change,
            "sd_change": self.sd_change,
            "source": self.source,
            "y_baseline": v(self.y_baseline),
            "sd_baseline": v(self.sd_baseline),
            "y_followup": v(self.y_followup),
            "sd_followup": v(self.sd_followup),
            "r_used": v(self.r_used),
            "responders": v(self.responders),
            "p_response": v(self.p_response),
            "q": v(self.q),
            "cutoff": v(self.cutoff),
        }


def as_int(v: object) -> int | None:
    if v is None or v == "NA":
        return None
    return int(v)


def as_float(v: object) -> float | None:
    if v is None or v == "NA":
        return None
    return float(v)


def pooled_sd_from_arms(arms: Iterable[tuple[int, float]]) -> float:
    arms = list(arms)
    df = sum(n for n, _ in arms) - len(arms)
    if df <= 0:
        raise ValueError("Invalid pooled SD degrees of freedom")
    pooled_var = sum((n - 1) * (sd**2) for n, sd in arms) / df
    if pooled_var <= 0:
        raise ValueError("Invalid pooled SD variance")
    return math.sqrt(pooled_var)


def find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, marker: str) -> int:
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "na[]" and ws.cell(r, 7).value == marker:
            return r
    raise ValueError(f"Could not find header row with marker {marker}")


def parse_cfb_block(ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int, end_row: int) -> list[ArmRow]:
    out: list[ArmRow] = []
    for r in range(start_row, end_row + 1):
        studyid = ws.cell(r, 26).value
        na = as_int(ws.cell(r, 1).value)
        if not studyid or studyid == "#" or na is None:
            continue

        arm_cells = []
        for arm in range(1, 6):
            trt = as_int(ws.cell(r, 1 + arm).value)
            n = as_int(ws.cell(r, 16 + arm).value)
            y = as_float(ws.cell(r, 6 + arm).value)
            sd = as_float(ws.cell(r, 11 + arm).value)
            if trt is None:
                continue
            if n is None or y is None or sd is None:
                raise ValueError(f"Missing CFB values in row {r}, arm {arm}")
            arm_cells.append((arm, trt, n, y, sd))

        pooled_sd = pooled_sd_from_arms((n, sd) for _, _, n, _, sd in arm_cells)
        for arm, trt, n, y, sd in arm_cells:
            out.append(
                ArmRow(
                    studyid=str(studyid),
                    na=na,
                    arm=arm,
                    treatment=trt,
                    n=n,
                    mean_change=y / pooled_sd,
                    sd_change=sd / pooled_sd,
                    source="smd",
                )
            )
    return out


def parse_bf_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
    rho: float,
) -> list[ArmRow]:
    out: list[ArmRow] = []
    for r in range(start_row, end_row + 1):
        studyid = ws.cell(r, 37).value
        na = as_int(ws.cell(r, 1).value)
        if not studyid or studyid == "#" or na is None:
            continue

        arm_cells = []
        for arm in range(1, 6):
            trt = as_int(ws.cell(r, 1 + arm).value)
            n = as_int(ws.cell(r, 26 + arm).value)
            yb = as_float(ws.cell(r, 6 + arm).value)
            sdb = as_float(ws.cell(r, 11 + arm).value)
            yf = as_float(ws.cell(r, 16 + arm).value)
            sdf = as_float(ws.cell(r, 21 + arm).value)
            if trt is None:
                continue
            if n is None or yb is None or sdb is None or yf is None or sdf is None:
                raise ValueError(f"Missing baseline/follow-up values in row {r}, arm {arm}")

            mean_change_raw = yf - yb
            var_change = (sdf**2) + (sdb**2) - (2 * rho * sdf * sdb)
            if var_change <= 0:
                raise ValueError(f"Non-positive change variance in row {r}, arm {arm}")
            sd_change_raw = math.sqrt(var_change)
            arm_cells.append((arm, trt, n, yb, sdb, yf, sdf, mean_change_raw, sd_change_raw))

        pooled_sd = pooled_sd_from_arms((n, sdb) for _, _, n, _, sdb, *_ in arm_cells)
        for arm, trt, n, yb, sdb, yf, sdf, mean_change_raw, sd_change_raw in arm_cells:
            out.append(
                ArmRow(
                    studyid=str(studyid),
                    na=na,
                    arm=arm,
                    treatment=trt,
                    n=n,
                    mean_change=mean_change_raw / pooled_sd,
                    sd_change=sd_change_raw / pooled_sd,
                    source="baseline_followup",
                    y_baseline=yb,
                    sd_baseline=sdb,
                    y_followup=yf,
                    sd_followup=sdf,
                    r_used=rho,
                )
            )
    return out


def parse_response_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
    rho: float,
    p_clip: float,
) -> list[ArmRow]:
    out: list[ArmRow] = []
    normal = NormalDist()

    for r in range(start_row, end_row + 1):
        studyid = ws.cell(r, 33).value
        na = as_int(ws.cell(r, 1).value)
        q = as_float(ws.cell(r, 27).value)
        if not studyid or studyid == "#" or na is None:
            continue
        if q is None:
            raise ValueError(f"Missing q in response row {r}")

        adj = math.sqrt(1 + (1 - q) * (1 - q - 2 * rho))
        if adj <= 0:
            raise ValueError(f"Invalid response adjustment in row {r}")

        arm_cells = []
        for arm in range(1, 6):
            trt = as_int(ws.cell(r, 1 + arm).value)
            responders = as_int(ws.cell(r, 6 + arm).value)
            n = as_int(ws.cell(r, 11 + arm).value)
            ybr = as_float(ws.cell(r, 16 + arm).value)
            sdbr = as_float(ws.cell(r, 21 + arm).value)
            if trt is None:
                continue
            if responders is None or n is None or ybr is None or sdbr is None:
                raise ValueError(f"Missing responder values in row {r}, arm {arm}")
            if not (0 <= responders <= n):
                raise ValueError(f"Invalid responders count in row {r}, arm {arm}")

            p = responders / n
            p_for_z = min(max(p, p_clip), 1 - p_clip)
            z = normal.inv_cdf(p_for_z)

            cutoff = -(q * ybr)
            mean_change_raw = -(q * ybr + (z * sdbr * adj))
            sd_change_raw = sdbr * adj

            arm_cells.append(
                (arm, trt, n, responders, p, ybr, sdbr, cutoff, mean_change_raw, sd_change_raw)
            )

        pooled_sd = pooled_sd_from_arms((n, sdbr) for _, _, n, _, _, sdbr, *_ in arm_cells)
        for arm, trt, n, responders, p, ybr, sdbr, cutoff, mean_change_raw, sd_change_raw in arm_cells:
            out.append(
                ArmRow(
                    studyid=str(studyid),
                    na=na,
                    arm=arm,
                    treatment=trt,
                    n=n,
                    mean_change=mean_change_raw / pooled_sd,
                    sd_change=sd_change_raw / pooled_sd,
                    source="responders",
                    y_baseline=ybr,
                    sd_baseline=sdbr,
                    responders=responders,
                    p_response=p,
                    q=q,
                    cutoff=cutoff,
                )
            )

    return out


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_xlsx", help="Path to mmc5.xlsx")
    parser.add_argument(
        "-o",
        "--output",
        default="combined_long_mean_change_dataset_ms_smd_bias_adj.csv",
        help="Output CSV path",
    )
    parser.add_argument(
        "--sheet",
        default="MS SMD bias-adj",
        help="Worksheet name",
    )
    parser.add_argument(
        "--rho",
        type=float,
        default=0.5,
        help="Baseline/follow-up correlation used in conversion formulas",
    )
    parser.add_argument(
        "--p-clip",
        type=float,
        default=1e-6,
        help="Clip probability for inverse normal transform in responder conversion",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input_xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{args.sheet}' not found")
    ws = wb[args.sheet]

    cfb_header = find_header_row(ws, "yCFB[,1]")
    bf_header = find_header_row(ws, "yB[,1]")
    resp_header = find_header_row(ws, "r[,1]")

    rows: list[ArmRow] = []
    rows.extend(parse_cfb_block(ws, cfb_header + 1, bf_header - 1))
    rows.extend(parse_bf_block(ws, bf_header + 1, resp_header - 1, rho=args.rho))
    rows.extend(parse_response_block(ws, resp_header + 1, ws.max_row, rho=args.rho, p_clip=args.p_clip))

    with open(args.output, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.to_csv_row())

    print(f"Wrote {len(rows)} rows to {args.output}")


if __name__ == "__main__":
    main()
